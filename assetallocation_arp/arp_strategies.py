import sys
import os

# excel based imports
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import assetallocation_arp.data_etl.import_data as gd
import assetallocation_arp.models.times as times
from assetallocation_arp.enum import models_names as models


def create_sheets_from_pandas_dataframe_input(workbook=None, sheet_name=None, column_value=None, list_df=None):
    if workbook is None or sheet_name is None or column_value is None or list_df is None:
        raise ("Invalid inputs")
    sheet_output = workbook.create_sheet(sheet_name)
    for df in list_df:
        for r in dataframe_to_rows(df, index=True, header=True):
            sheet_output.append(r)


def create_sheets_from_pandas_dataframe(workbook=None, sheet_name=None, column_value=None, df=None):
    if workbook is None or sheet_name is None or column_value is None or df is None:
        raise ("Invalid inputs")
    sheet_output = workbook.create_sheet(sheet_name)
    sheet_output['A1'] = column_value
    for r in dataframe_to_rows(df, index=True, header=True):
        sheet_output.append(r)


def run_model(model_type, mat_file=None, input_file=None):
    if model_type == models.Models.times.name:
        # get inputs from excel and matlab data
        times_inputs, asset_inputs, all_data = gd.extract_inputs_and_mat_data(model_type, mat_file, input_file)
        # run strategy
        signals, returns, r, positioning = times.format_data_and_calc(times_inputs, asset_inputs, all_data)
        # write results to output sheet
        write_output_to_excel({models.Models.times.name: (asset_inputs, positioning, r, signals, times_inputs)})

    if model_type == models.Models.maven.name:
        print(model_type)
    if model_type == models.Models.effect.name:
        print(model_type)
    if model_type == models.Models.curp.name:
        print(model_type)
    if model_type == models.Models.fica.name:
        print(model_type)
    if model_type == models.Models.factor.name:
        print(model_type)
    if model_type == models.Models.comca.name:
        print(model_type)


def write_output_to_excel(model_outputs):
    if models.Models.times.name in model_outputs.keys():
        asset_inputs, positioning, returns, signals, times_inputs = model_outputs[models.Models.times.name]
        path = os.path.join(os.path.dirname(__file__), "times_model.xls")
    wb = Workbook()
    create_sheets_from_pandas_dataframe(wb, 'signal', "TIMES Signals", signals)
    create_sheets_from_pandas_dataframe(wb, 'returns', "TIMES Returns", returns)
    create_sheets_from_pandas_dataframe(wb, 'positioning', "TIMES Positioning", positioning)
    create_sheets_from_pandas_dataframe_input(wb, 'times_inputs', "TIMES Inputs", [asset_inputs, times_inputs])

    wb.save(path)


def get_inputs_from_excel():
    # select data from excel

    mat_file = xw.Range('rng_mat_file_path').value

    model_type = xw.Range('rng_model_type').value

    # run selected model

    run_model(model_type, mat_file, xw.Book.caller().fullname)


def get_inputs_from_python(model):
    # launch the script from Python
    mat_file = None
    input_file = None
    models_list = [model.name for model in models.Models]

    if model in models_list:
        model_type = model
        run_model(model_type, mat_file, input_file)
    else:
        raise NameError("Your input is incorrect.")


def get_input_user():
    model_str = input("Choose a Model: ")
    return model_str


if __name__ == "__main__":
    sys.exit(get_inputs_from_python(get_input_user()))
